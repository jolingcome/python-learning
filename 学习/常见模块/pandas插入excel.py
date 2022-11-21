#pandas是原有的excel中插入数据，并不覆盖.
# 如果直接覆盖则用pandas的df.to_excel(filename)
import pandas as pd
from openpyxl import load_workbook
import xlwt

#将dict类型的数据追加写入到现有的Excel中
def write_to_exist_excel(filename,sheetname):
    df_old=pd.DataFrame(pd.read_excel(filename,sheet_name=sheetname))
    row_old=df_old.shape[0] #获取原数据的行数

    data_added={"a": 99, "b": 98, "c": 97}
    df=pd.DataFrame(data_added,index=[0])  # 如果dict的value只有一行，加上index = [0]即可

    book=load_workbook(filename)
    writer=pd.ExcelWriter(filename,engine='openpyxl')
    writer.book=book
    writer.sheets=dict((ws.title,ws) for ws in book.worksheets)

    # 将data_added数据写入Excel中
    df.to_excel(writer, sheet_name=sheetname, startrow=row_old + 1, index=False, header=False)

    writer.save()  # 保存


# 将list[dict]类型的数据追加写入到现有的Excel中
def write_to_exist_excel2(fileName, sheetName):
    df_old = pd.DataFrame(pd.read_excel(fileName, sheet_name=sheetName))  # 读取原数据文件和表
    row_old = df_old.shape[0]  # 获取原数据的行数

    data_added = [{"a": 9, "b": 8, "c": 7, "d": 6},
                  {"a": 37, "b": 38, "c": 39, "d": 40}]
    df = pd.DataFrame(data_added)

    book = load_workbook(fileName)
    writer = pd.ExcelWriter(fileName, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # 将data_added数据写入Excel中
    df.to_excel(writer, sheet_name=sheetName, startrow=row_old + 1, index=False, header=False)

    writer.save()  # 保存


# 3、调用 数据存储至excel中,写入新的excel表中,可以用pandas直接写入
# def excel_data(excel_name,testdata):
#     # a、创建表格，初始化
#     workbook = xlwt.Workbook(encoding='utf-8')
#     data_sheet = workbook.add_sheet('sheet1')
#     # b、输入表头
#     excel_title = ['用例编号','所属模块','子模块', '优先级','前置条件', '用例步骤', '预期结果']
#     for i in range(len(excel_title)):
#         data_sheet.write(0, i, excel_title[i])
#     #c、输入case数据
#     for x in range(len(testdata)):
#         for y in range(len(excel_title)):
#             data_sheet.write(x+1, y, ng_test(testdata)[x][y])
#     # d、保存
#     workbook.save(excel_name)
#     return data_sheet

excel_path=r"D:\studio\测试用例\SFE3.0测试用例模板.xlsx"
write_to_exist_excel(excel_path,sheetname="授权")
